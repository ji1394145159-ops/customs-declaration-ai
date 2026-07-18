import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from typing import Dict


def _find_chinese_font() -> str:
    """查找可用的中文字体"""
    font_paths = [
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/msyh.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
    ]
    # 遍历 noto-cjk 目录查找字体
    noto_dir = "/usr/share/fonts/opentype/noto"
    if os.path.isdir(noto_dir):
        for f in os.listdir(noto_dir):
            if "CJK" in f and f.endswith((".ttc", ".ttf")):
                font_paths.append(os.path.join(noto_dir, f))
    for path in font_paths:
        if os.path.exists(path):
            return path
    return None


def export_to_excel(declaration_data: dict) -> bytes:
    """导出申报单为Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "报关申报单"

    # 标题样式
    title_font = Font(name="微软雅黑", size=14, bold=True)
    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells("A1:D1")
    ws["A1"] = "跨境电商报关申报单"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # 基本信息
    ws["A2"] = f"目标国家: {declaration_data.get('target_country', '')}"
    ws["A3"] = f"HS编码: {declaration_data.get('hs_code', '')}"
    ws["A4"] = f"完整度评分: {declaration_data.get('completeness_score', '')}"

    # 申报要素表格
    row = 6
    headers = ["序号", "申报要素", "内容", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    elements = declaration_data.get("declaration_elements", {})
    risk_fields = {r["field"] for r in declaration_data.get("risk_fields", [])}

    for idx, (key, value) in enumerate(elements.items(), 1):
        row += 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=key).border = thin_border
        ws.cell(row=row, column=3, value=str(value) if value else "").border = thin_border
        note = "⚠️ 需人工确认" if key in risk_fields else ""
        ws.cell(row=row, column=4, value=note).border = thin_border

    # 合规提示
    row += 2
    ws.cell(row=row, column=1, value="合规提示:").font = header_font
    for note in declaration_data.get("compliance_notes", []):
        row += 1
        ws.cell(row=row, column=1, value=f"• {note}")

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_to_pdf(declaration_data: dict) -> bytes:
    """导出申报单为PDF"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # 尝试注册中文字体
    font_name = "Helvetica"
    chinese_font_path = _find_chinese_font()
    if chinese_font_path:
        try:
            pdfmetrics.registerFont(TTFont("ChineseFont", chinese_font_path))
            font_name = "ChineseFont"
        except Exception:
            pass

    # 标题
    c.setFont(font_name, 18)
    c.drawCentredString(width / 2, height - 40 * mm, "Cross-border E-commerce Customs Declaration")

    c.setFont(font_name, 14)
    c.drawCentredString(width / 2, height - 50 * mm, "跨境电商报关申报单")

    # 基本信息
    y = height - 70 * mm
    c.setFont(font_name, 11)
    c.drawString(25 * mm, y, f"Target Country / 目标国家: {declaration_data.get('target_country', '')}")
    y -= 8 * mm
    c.drawString(25 * mm, y, f"HS Code / HS编码: {declaration_data.get('hs_code', '')}")
    y -= 8 * mm
    c.drawString(25 * mm, y, f"Completeness / 完整度: {declaration_data.get('completeness_score', '')}%")

    # 申报要素
    y -= 15 * mm
    c.setFont(font_name, 13)
    c.drawString(25 * mm, y, "Declaration Elements / 申报要素:")
    y -= 10 * mm

    c.setFont(font_name, 10)
    elements = declaration_data.get("declaration_elements", {})
    risk_fields = {r["field"] for r in declaration_data.get("risk_fields", [])}

    for key, value in elements.items():
        if y < 30 * mm:
            c.showPage()
            y = height - 30 * mm
            c.setFont(font_name, 10)

        display_value = str(value) if value else "N/A"
        risk_mark = " [!]" if key in risk_fields else ""
        c.drawString(25 * mm, y, f"{key}: {display_value}{risk_mark}")
        y -= 6 * mm

    # 合规提示
    y -= 10 * mm
    if y < 50 * mm:
        c.showPage()
        y = height - 30 * mm

    c.setFont(font_name, 13)
    c.drawString(25 * mm, y, "Compliance Notes / 合规提示:")
    y -= 8 * mm
    c.setFont(font_name, 10)
    for note in declaration_data.get("compliance_notes", []):
        if y < 30 * mm:
            c.showPage()
            y = height - 30 * mm
            c.setFont(font_name, 10)
        c.drawString(25 * mm, y, f"• {note}")
        y -= 6 * mm

    c.save()
    return buffer.getvalue()
